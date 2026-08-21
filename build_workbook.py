import openpyxl
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ---------- Styles ----------
FONT = "Arial"
TONIK_PURPLE = "785AFF"      # Tonik primary brand color (Cornflower Blue / purple)
TONIK_DARK = "682763"        # Tonik secondary brand color ("Finn")
TONIK_TINT = "EDE8FF"        # light tint of the primary purple, for section bands

TITLE_FONT = Font(name=FONT, size=14, bold=True, color=TONIK_DARK)
HEADER_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
LABEL_FONT = Font(name=FONT, size=10, bold=True)
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="666666")
INPUT_FONT = Font(name=FONT, size=10, bold=True, color="0000FF")
FORMULA_FONT = Font(name=FONT, size=10, color="000000")
LINK_FONT = Font(name=FONT, size=10, color="008000")

HEADER_FILL = PatternFill("solid", fgColor=TONIK_PURPLE)
SECTION_FILL = PatternFill("solid", fgColor=TONIK_TINT)
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
TOTAL_FILL = PatternFill("solid", fgColor="E8E8E8")
RESULT_FILL = PatternFill("solid", fgColor="C6E0B4")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CUR = '#,##0.00;(#,##0.00);"-"'
PCT2 = '0.00%'
DATEFMT = 'mmmm d, yyyy'

def style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_block(ws, text, span=8, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")

def note(ws, row, col, text, span=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = NOTE_FONT
    if span:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    return c

# =========================================================================
# SHEET 0: INSTRUCTIONS
# =========================================================================
ws = wb.active
ws.title = "Instructions"
set_col_widths(ws, [4, 100])
title_block(ws, "Loan Restructuring Calculator — How to Use", span=2)
ws.row_dimensions[1].height = 24

rows = [
    ("", ""),
    ("Purpose", "Encodes an existing loan's details, lets you enter the borrower's last payment date and days-past-due (DPD), "
                "and automatically computes the restructured loan — with the basis for every figure shown."),
    ("", ""),
    ("Color legend", ""),
    ("  Blue bold text on yellow fill", "Means: input cell. This is the ONLY thing you should type into."),
    ("  Black text", "Means: formula. Do not overwrite — it recalculates automatically."),
    ("  Green text", "Means: a value linked in from another sheet."),
    ("", ""),
    ("Sheet order", ""),
    ("  1. Original Loan Setup", "Encode the existing loan's terms (principal, rate, term, fees) exactly as shown on its Promissory Note."),
    ("  2. Original Amortization Schedule", "Auto-generated from Sheet 1 — mirrors the bank's amortization schedule format."),
    ("  3. Restructuring Workup", "Enter the Last Payment Date and DPD here. Everything else on this sheet — outstanding principal, "
                                    "accrued interest, penalty, and the new restructured principal — is calculated, with its basis stated "
                                    "next to each figure. Policy: DPD > 90 days -> restructured principal = Outstanding Principal only. "
                                    "DPD <= 90 days -> restructured principal = Outstanding Principal + Accrued Interest. (Penalty "
                                    "capitalization is a separate manual Yes/No toggle.)"),
    ("  4. Restructured Amortization Schedule", "Auto-generated new schedule for the restructured loan, same mechanics as Sheet 2."),
    ("  5. Summary", "Side-by-side comparison of the original vs. restructured loan for quick reference."),
    ("", ""),
    ("Sample scenario", "This workbook ships pre-filled with an illustrative unsecured cash loan scenario "
                          "(₱20,000 principal, 95% APR, 12 installments, restructured into a 47.5% APR, 16-installment "
                          "loan) so you can see how the mechanics work end-to-end. No client or borrower information "
                          "is stored in this file. Overwrite the yellow cells on Sheets 1 and 3 with a real loan's "
                          "terms to use the calculator."),
]
r = 2
for label, text in rows:
    if label and not text:
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = Font(name=FONT, size=11, bold=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    else:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        c2 = ws.cell(row=r, column=2, value=text)
        c2.font = Font(name=FONT, size=10)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30 if text else 14
    r += 1

print("Instructions sheet built")

# =========================================================================
# SHEET 1: ORIGINAL LOAN SETUP
# =========================================================================
ws = wb.create_sheet("1. Original Loan Setup")
set_col_widths(ws, [3, 34, 20, 3, 46, 20])
title_block(ws, "Original Loan — Encoded Details", span=6)
ws.row_dimensions[1].height = 24
note(ws, 2, 2, "Source: Tonik Digital Bank Promissory Note. Yellow cells are inputs — replace with a new borrower's terms to reuse this workbook.", span=5)

def field(ws, row, label, value, fmt=None, col_label=2, col_val=3, is_input=True, font=None):
    lc = ws.cell(row=row, column=col_label, value=label)
    lc.font = LABEL_FONT
    vc = ws.cell(row=row, column=col_val, value=value)
    vc.font = font if font else (INPUT_FONT if is_input else FORMULA_FONT)
    if is_input:
        vc.fill = INPUT_FILL
    vc.border = BORDER
    if fmt:
        vc.number_format = fmt
    vc.alignment = Alignment(horizontal="left")
    return vc

r = 4
ws.cell(row=r, column=2, value="LOAN IDENTIFICATION").font = LABEL_FONT
ws.cell(row=r, column=2).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1
field(ws, r, "Loan Reference / ID (optional, internal use)", "e.g. LN-000001"); R_ACCTNO=r; r += 1

r += 1
ws.cell(row=r, column=2, value="ORIGINAL LOAN TERMS (per Promissory Note)").font = LABEL_FONT
ws.cell(row=r, column=2).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1
field(ws, r, "Principal Amount (PHP)", 20000.00, CUR); R_PRIN=r; r += 1
field(ws, r, "Date Granted", datetime.date(2025,11,30), DATEFMT); R_GRANT=r; r += 1
field(ws, r, "Annual Interest Rate (APR)", 0.95, PCT2); R_APR=r; r += 1
vc = field(ws, r, "Monthly Interest Rate", f"=C{R_APR}/12", PCT2, is_input=False); R_MRATE=r; r += 1
field(ws, r, "Number of Installments", 12); R_NPER=r; r += 1
field(ws, r, "First Monthly Due Date", datetime.date(2026,1,2), DATEFMT); R_FIRSTDUE=r; r += 1
field(ws, r, "Processing Fee (PHP)", 500.00, CUR); R_PROCFEE=r; r += 1
field(ws, r, "Documentary Stamps Tax (PHP)", 150.00, CUR); R_DST=r; r += 1
vc = field(ws, r, "Net Proceeds of Loan (PHP)", f"=C{R_PRIN}-C{R_PROCFEE}-C{R_DST}", CUR, is_input=False); R_NETPROC=r; r += 1
field(ws, r, "Value-Added Services (VAS) Fee %", 0.0911, PCT2); R_VASPCT=r; r += 1
note(ws, r, 2, "VAS fee = this % applied to each installment's (principal + interest).", span=2); r += 1
field(ws, r, "Late Payment Fee per missed installment (PHP)", 500.00, CUR); R_LATEFEE=r; r += 1

r += 1
ws.cell(row=r, column=2, value="COMPUTED SUMMARY (for reference)").font = LABEL_FONT
ws.cell(row=r, column=2).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1
field(ws, r, "Fixed Monthly Installment (Principal+Interest)",
      f"=-PMT(C{R_MRATE},C{R_NPER},C{R_PRIN})", CUR, is_input=False); R_PMT=r; r += 1
field(ws, r, "Maturity Date", f"=EDATE(C{R_FIRSTDUE},C{R_NPER}-1)", DATEFMT, is_input=False); R_MATURITY=r; r += 1

# store row map for cross-sheet refs
SETUP = dict(ACCTNO=R_ACCTNO, PRIN=R_PRIN, GRANT=R_GRANT, APR=R_APR,
             MRATE=R_MRATE, NPER=R_NPER, FIRSTDUE=R_FIRSTDUE, PROCFEE=R_PROCFEE, DST=R_DST,
             NETPROC=R_NETPROC, VASPCT=R_VASPCT, LATEFEE=R_LATEFEE, PMT=R_PMT, MATURITY=R_MATURITY)
print("Setup rows:", SETUP)

print("Sheet 1 built")

# =========================================================================
# Generic amortization-schedule builder (used for Sheet 2 and Sheet 4)
# =========================================================================
def build_amort_sheet(ws, title_text, prin_ref, mrate_ref, nper_ref, firstdue_ref,
                       apr_ref, pmt_ref, grant_ref, vaspct_ref=None, latefee_ref=None, max_rows=20,
                       source_note=""):
    """prin_ref/mrate_ref/... are fully-qualified formula fragments, e.g. \"'1. Original Loan Setup'!$C$10\" """
    set_col_widths(ws, [4, 16, 15, 15, 15, 14, 15, 16] if vaspct_ref else [4, 16, 15, 15, 15, 15, 16])
    title_block(ws, title_text, span=8)
    ws.row_dimensions[1].height = 22
    if source_note:
        note(ws, 2, 1, source_note, span=8)

    hdr_row = 4
    if vaspct_ref:
        headers = ["No.", "Installment Date", "Beginning\nBalance", "Principal", "Interest", "VAS Fee", "Total\nPayment", "Ending\nBalance"]
    else:
        headers = ["No.", "Installment Date", "Beginning\nBalance", "Principal", "Interest", "Total\nPayment", "Ending\nBalance"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hdr_row, column=i, value=h)
    style_header_row(ws, hdr_row, 1, len(headers))
    ws.row_dimensions[hdr_row].height = 30

    first_data_row = hdr_row + 1
    ncols = len(headers)
    for i in range(max_rows):
        row = first_data_row + i
        n = i + 1  # installment number
        c_no, c_date, c_beg, c_prin, c_int = 1, 2, 3, 4, 5
        if vaspct_ref:
            c_vas, c_tot, c_end = 6, 7, 8
        else:
            c_tot, c_end = 6, 7

        ws.cell(row=row, column=c_no, value=f"=IF({n}<={nper_ref},{n},\"\")").font = FORMULA_FONT

        # date
        ws.cell(row=row, column=c_date,
                value=f"=IF({n}<={nper_ref},EDATE({firstdue_ref},{n}-1),\"\")").number_format = DATEFMT
        ws.cell(row=row, column=c_date).font = FORMULA_FONT

        # beginning balance
        if i == 0:
            beg_formula = f"=IF({n}<={nper_ref},{prin_ref},\"\")"
        else:
            prev_end_addr = f"{get_column_letter(c_end)}{row-1}"
            beg_formula = f"=IF({n}<={nper_ref},{prev_end_addr},\"\")"
        ws.cell(row=row, column=c_beg, value=beg_formula).number_format = CUR
        ws.cell(row=row, column=c_beg).font = FORMULA_FONT

        beg_addr = f"{get_column_letter(c_beg)}{row}"
        date_addr = f"{get_column_letter(c_date)}{row}"
        # interest: actual/365 daily accrual on the beginning balance over the actual number of
        # days in the one-month billing cycle ending on this installment's date (i.e. days since
        # EDATE(this date, -1) — 28/30/31 depending on the calendar month), matching the bank's schedule
        ws.cell(row=row, column=c_int,
                value=f"=IF({n}<={nper_ref},{beg_addr}*{apr_ref}/365*({date_addr}-EDATE({date_addr},-1)),\"\")").number_format = CUR
        ws.cell(row=row, column=c_int).font = FORMULA_FONT
        int_addr = f"{get_column_letter(c_int)}{row}"

        # principal: last installment forces payoff of beginning balance; else PMT - interest
        prin_formula = (f"=IF({n}<{nper_ref},{pmt_ref}-{int_addr},"
                         f"IF({n}={nper_ref},{beg_addr},\"\"))")
        ws.cell(row=row, column=c_prin, value=prin_formula).number_format = CUR
        ws.cell(row=row, column=c_prin).font = FORMULA_FONT
        prin_addr = f"{get_column_letter(c_prin)}{row}"

        # VAS fee (original loan only)
        if vaspct_ref:
            vas_formula = f"=IF({n}<={nper_ref},({prin_addr}+{int_addr})*{vaspct_ref},\"\")"
            ws.cell(row=row, column=c_vas, value=vas_formula).number_format = CUR
            ws.cell(row=row, column=c_vas).font = FORMULA_FONT
            vas_addr = f"{get_column_letter(c_vas)}{row}"
            tot_formula = f"=IF({n}<={nper_ref},{prin_addr}+{int_addr}+{vas_addr},\"\")"
        else:
            tot_formula = f"=IF({n}<={nper_ref},{prin_addr}+{int_addr},\"\")"
        ws.cell(row=row, column=c_tot, value=tot_formula).number_format = CUR
        ws.cell(row=row, column=c_tot).font = FORMULA_FONT

        # ending balance
        end_formula = f"=IF({n}<={nper_ref},{beg_addr}-{prin_addr},\"\")"
        ws.cell(row=row, column=c_end, value=end_formula).number_format = CUR
        ws.cell(row=row, column=c_end).font = FORMULA_FONT

        for cc in range(1, ncols + 1):
            ws.cell(row=row, column=cc).border = BORDER
            ws.cell(row=row, column=cc).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=c_no).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=c_date).alignment = Alignment(horizontal="center")

    total_row = first_data_row + max_rows
    ws.cell(row=total_row, column=2, value="TOTAL").font = LABEL_FONT
    ws.cell(row=total_row, column=2).fill = TOTAL_FILL
    for cc, col_letter_idx in [(c_prin, c_prin), (c_int, c_int)] if False else []:
        pass
    sum_cols = [c_prin, c_int] + ([c_vas] if vaspct_ref else []) + [c_tot]
    for cc in sum_cols:
        col_letter = get_column_letter(cc)
        f = f"=SUM({col_letter}{first_data_row}:{col_letter}{total_row-1})"
        cell = ws.cell(row=total_row, column=cc, value=f)
        cell.number_format = CUR
        cell.font = LABEL_FONT
        cell.fill = TOTAL_FILL
        cell.border = BORDER
    for cc in range(1, ncols + 1):
        ws.cell(row=total_row, column=cc).fill = TOTAL_FILL
        ws.cell(row=total_row, column=cc).border = BORDER
    return dict(first_data_row=first_data_row, total_row=total_row, hdr_row=hdr_row,
                c_end=c_end, c_date=c_date, c_no=c_no)

# ---- Sheet 2: Original Amortization Schedule ----
ws2 = wb.create_sheet("2. Original Amort Schedule")
SETUP_SHEET = "'1. Original Loan Setup'"
refs = dict(
    prin_ref=f"{SETUP_SHEET}!$C${SETUP['PRIN']}",
    mrate_ref=f"{SETUP_SHEET}!$C${SETUP['MRATE']}",
    nper_ref=f"{SETUP_SHEET}!$C${SETUP['NPER']}",
    firstdue_ref=f"{SETUP_SHEET}!$C${SETUP['FIRSTDUE']}",
    apr_ref=f"{SETUP_SHEET}!$C${SETUP['APR']}",
    pmt_ref=f"{SETUP_SHEET}!$C${SETUP['PMT']}",
    grant_ref=f"{SETUP_SHEET}!$C${SETUP['GRANT']}",
    vaspct_ref=f"{SETUP_SHEET}!$C${SETUP['VASPCT']}",
)
LOC2 = build_amort_sheet(ws2, "Original Loan — Amortization Schedule", max_rows=20,
                          source_note="Auto-generated from Sheet 1. Installment = fixed annuity payment at the Monthly "
                                      "Rate (Principal+Interest constant every period, via the PMT function). "
                                      "Interest = Beginning Balance x (Annual Rate/365) x actual days in the billing "
                                      "cycle (28/30/31, per the calendar month ending on each installment date) — this "
                                      "reproduces the source Promissory Note's schedule to the peso for installments "
                                      "1-11. Principal = Installment - Interest, except the FINAL installment, which "
                                      "pays off the remaining balance exactly; banks typically true up the final "
                                      "installment's interest so total interest matches the amount disclosed at "
                                      "origination, so this model's last-row interest can differ slightly from the "
                                      "bank's own statement (immaterial to the outstanding-principal figures the "
                                      "restructuring calculation relies on). VAS Fee = 9.11% of (Principal+Interest).",
                          **refs)
print("Sheet 2 built:", LOC2)

# =========================================================================
# SHEET 3: RESTRUCTURING WORKUP
# =========================================================================
ws3 = wb.create_sheet("3. Restructuring Workup")
set_col_widths(ws3, [3, 42, 20, 3, 55])
title_block(ws3, "Restructuring Workup", span=5)
ws3.row_dimensions[1].height = 24
note(ws3, 2, 2, "Type only into the yellow cells (Last Payment Date and DPD). Every other figure below is computed, "
                 "with its basis stated in column E.", span=4)

AMORT2_SHEET = "'2. Original Amort Schedule'"
end_col_letter = get_column_letter(LOC2['c_end'])
date_col_letter = get_column_letter(LOC2['c_date'])
fdr = LOC2['first_data_row']
tot = LOC2['total_row']

r = 4
ws3.cell(row=r, column=2, value="STEP 1 — COLLECTIONS INPUT (from loan monitoring system)").font = LABEL_FONT
ws3.cell(row=r, column=2).fill = SECTION_FILL
ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1
R_LASTPAY = field(ws3, r, "Last Payment Date (last installment date actually paid)", datetime.date(2026,3,2), DATEFMT); R_LASTPAY=r
note(ws3, r, 5, "Must match an Installment Date on Sheet 2. Leave blank / use the Date Granted if no installment has been paid yet.")
r += 1
R_ASOF = r
field(ws3, r, "As-of / Report Date", datetime.date(2026,8,18), DATEFMT); r += 1
R_DPD = r
field(ws3, r, "Days Past Due (DPD) — as reported by collections", 138); r += 1

r += 1
ws3.cell(row=r, column=2, value="STEP 2 — OUTSTANDING EXPOSURE AT DEFAULT (computed)").font = LABEL_FONT
ws3.cell(row=r, column=2).fill = SECTION_FILL
ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1

# Outstanding principal: lookup ending balance on row where date = Last Payment Date
R_OSPRIN = r
f = (f"=IFERROR(INDEX({AMORT2_SHEET}!${end_col_letter}${fdr}:${end_col_letter}${tot-1},"
     f"MATCH(C{R_LASTPAY},{AMORT2_SHEET}!${date_col_letter}${fdr}:${date_col_letter}${tot-1},0)),"
     f"{SETUP_SHEET}!$C${SETUP['PRIN']})")
field(ws3, r, "Outstanding Principal (PHP)", f, CUR, is_input=False)
note(ws3, r, 5, "Basis: Ending Balance on Sheet 2 for the row whose Installment Date = Last Payment Date above "
                "(i.e., the balance left after the last installment the borrower actually paid).")
r += 1

R_ACCINT = r
f = f"=C{R_OSPRIN}*{SETUP_SHEET}!$C${SETUP['APR']}/365*C{R_DPD}"
field(ws3, r, "Accrued Interest during delinquency (PHP)", f, CUR, is_input=False)
note(ws3, r, 5, "Basis: Outstanding Principal x Original Contractual Annual Rate / 365 x DPD (simple daily interest "
                "per the original Promissory Note's stated rate). Adjust the rate assumption in Sheet 1 if the bank "
                "applies a different penalty-interest rate during delinquency.")
r += 1

R_PENALTY = r
f = f"=ROUNDUP(C{R_DPD}/30,0)*{SETUP_SHEET}!$C${SETUP['LATEFEE']}"
field(ws3, r, "Penalty / Late Payment Fees (PHP)", f, CUR, is_input=False)
note(ws3, r, 5, "Basis: Late Payment Fee stated on the original Promissory Note x number of missed 30-day installment "
                "cycles (ROUNDUP(DPD/30)).")
r += 1

R_TOTEXP = r
f = f"=C{R_OSPRIN}+C{R_ACCINT}+C{R_PENALTY}"
field(ws3, r, "Total Outstanding Exposure (PHP)", f, CUR, is_input=False)
note(ws3, r, 5, "Basis: Outstanding Principal + Accrued Interest + Penalty. This is the borrower's full exposure as of the "
                "As-of Date, before any restructuring relief.")
r += 1

r += 1
ws3.cell(row=r, column=2, value="STEP 3 — RESTRUCTURING DECISION (per DPD policy; Penalty toggle is manual)").font = LABEL_FONT
ws3.cell(row=r, column=2).fill = SECTION_FILL
ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1

R_INCINT = r
f = f'=IF(C{R_DPD}>90,"No","Yes")'
field(ws3, r, "Capitalize Accrued Interest into new Principal? (auto, per DPD policy)", f, is_input=False)
note(ws3, r, 5, "Policy rule (not manually editable): if DPD > 90 days, the restructured loan is the Outstanding "
                "Principal ONLY (accrued interest is waived). If DPD is 90 days or less, the restructured loan is "
                "Outstanding Principal PLUS Accrued Interest (interest is capitalized). Driven automatically by the "
                "DPD entered in Step 1.")
r += 1
R_INCPEN = r
field(ws3, r, "Capitalize Penalty into new Principal? (manual)", "No")
note(ws3, r, 5, "Yes/No — manual choice, not covered by the DPD policy above. In the sample restructured PN "
                "supplied, penalty was WAIVED (No).")
r += 1

dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
ws3.add_data_validation(dv)
dv.add(ws3.cell(row=R_INCPEN, column=3))

r += 1
ws3.cell(row=r, column=2, value="RESTRUCTURED LOAN — NEW TERMS").font = LABEL_FONT
ws3.cell(row=r, column=2).fill = SECTION_FILL
ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1

R_NEWPRIN = r
f = (f"=C{R_OSPRIN}+IF(C{R_INCINT}=\"Yes\",C{R_ACCINT},0)+IF(C{R_INCPEN}=\"Yes\",C{R_PENALTY},0)")
field(ws3, r, "New Restructured Principal (PHP)", f, CUR, is_input=False)
ws3.cell(row=r, column=3).fill = RESULT_FILL
note(ws3, r, 5, "Basis: Outstanding Principal, plus Accrued Interest if DPD <= 90 (per policy above), plus Penalty "
                "only if manually capitalized.")
r += 1
R_NEWAPR = r
field(ws3, r, "New Annual Interest Rate (APR)", 0.475, PCT2)
r += 1
R_NEWMRATE = r
field(ws3, r, "New Monthly Interest Rate", f"=C{R_NEWAPR}/12", PCT2, is_input=False)
r += 1
R_NEWNPER = r
field(ws3, r, "New Number of Installments", 16)
r += 1
R_NEWFIRSTDUE = r
field(ws3, r, "New First Monthly Due Date", datetime.date(2026,10,5), DATEFMT)
r += 1
R_NEWGRANT = r
field(ws3, r, "New Date Granted (restructuring date)", datetime.date(2026,8,18), DATEFMT)
r += 1
R_NEWPMT = r
f = f"=-PMT(C{R_NEWMRATE},C{R_NEWNPER},C{R_NEWPRIN})"
field(ws3, r, "New Fixed Monthly Installment (Principal+Interest)", f, CUR, is_input=False)
r += 1
R_NEWMAT = r
f = f"=EDATE(C{R_NEWFIRSTDUE},C{R_NEWNPER}-1)"
field(ws3, r, "New Maturity Date", f, DATEFMT, is_input=False)
r += 1
R_TOTALRESTRUCT = r
f = f"=C{R_NEWPMT}*C{R_NEWNPER}"
field(ws3, r, "Total of New Payments over the term (PHP)", f, CUR, is_input=False)
ws3.cell(row=r, column=3).fill = RESULT_FILL
r += 1
R_WAIVED = r
f = (f"=IF(C{R_INCINT}=\"No\",C{R_ACCINT},0)+IF(C{R_INCPEN}=\"No\",C{R_PENALTY},0)")
field(ws3, r, "Total Waived (Interest + Penalty not capitalized) (PHP)", f, CUR, is_input=False)

WORKUP = dict(LASTPAY=R_LASTPAY, ASOF=R_ASOF, DPD=R_DPD, OSPRIN=R_OSPRIN, ACCINT=R_ACCINT,
              PENALTY=R_PENALTY, TOTEXP=R_TOTEXP, INCINT=R_INCINT, INCPEN=R_INCPEN,
              NEWPRIN=R_NEWPRIN, NEWAPR=R_NEWAPR, NEWMRATE=R_NEWMRATE, NEWNPER=R_NEWNPER,
              NEWFIRSTDUE=R_NEWFIRSTDUE, NEWGRANT=R_NEWGRANT, NEWPMT=R_NEWPMT, NEWMAT=R_NEWMAT,
              TOTALRESTRUCT=R_TOTALRESTRUCT, WAIVED=R_WAIVED)
print("Workup rows:", WORKUP)

# =========================================================================
# SHEET 4: RESTRUCTURED AMORTIZATION SCHEDULE
# =========================================================================
ws4 = wb.create_sheet("4. Restructured Amort Sched")
WORKUP_SHEET = "'3. Restructuring Workup'"
refs4 = dict(
    prin_ref=f"{WORKUP_SHEET}!$C${WORKUP['NEWPRIN']}",
    mrate_ref=f"{WORKUP_SHEET}!$C${WORKUP['NEWMRATE']}",
    nper_ref=f"{WORKUP_SHEET}!$C${WORKUP['NEWNPER']}",
    firstdue_ref=f"{WORKUP_SHEET}!$C${WORKUP['NEWFIRSTDUE']}",
    apr_ref=f"{WORKUP_SHEET}!$C${WORKUP['NEWAPR']}",
    pmt_ref=f"{WORKUP_SHEET}!$C${WORKUP['NEWPMT']}",
    grant_ref=f"{WORKUP_SHEET}!$C${WORKUP['NEWGRANT']}",
    vaspct_ref=None,
)
LOC4 = build_amort_sheet(ws4, "Restructured Loan — Amortization Schedule", max_rows=24,
                          source_note="Auto-generated from Sheet 3 (New Restructured Principal, Rate, Term, First Due "
                                      "Date, Date Granted). Same actual/365 daily-interest mechanics as Sheet 2; no "
                                      "VAS fee on the restructured loan (per the sample restructured Promissory Note "
                                      "supplied).",
                          **refs4)
print("Sheet 4 built:", LOC4)

# =========================================================================
# SHEET 5: SUMMARY
# =========================================================================
ws5 = wb.create_sheet("5. Summary")
set_col_widths(ws5, [3, 34, 20, 22, 3])
title_block(ws5, "Original vs. Restructured Loan — Summary", span=4)
ws5.row_dimensions[1].height = 24
note(ws5, 2, 2, "All cells below are linked from Sheets 1–4. Nothing to type here.", span=3)

r = 4
headers = ["", "Original Loan", "Restructured Loan"]
for i, h in enumerate(headers, start=2):
    c = ws5.cell(row=r, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
r += 1

def sumrow(ws, r, label, orig_formula, new_formula, fmt=CUR):
    ws.cell(row=r, column=2, value=label).font = LABEL_FONT
    c1 = ws.cell(row=r, column=3, value=orig_formula)
    c2 = ws.cell(row=r, column=4, value=new_formula)
    for c in (c1, c2):
        c.font = LINK_FONT
        c.number_format = fmt
        c.border = BORDER
        c.alignment = Alignment(horizontal="right")
    ws.cell(row=r, column=2).border = BORDER
    return r + 1

ws5.cell(row=r, column=2, value="Loan Reference / ID").font = LABEL_FONT
ws5.cell(row=r, column=2).border = BORDER
c1 = ws5.cell(row=r, column=3, value=f"={SETUP_SHEET}!C{SETUP['ACCTNO']}")
c1.font = LINK_FONT; c1.border = BORDER
c2 = ws5.cell(row=r, column=4, value="New reference assigned upon restructuring")
c2.font = NOTE_FONT; c2.border = BORDER
r += 1
r = sumrow(ws5, r, "Principal (PHP)", f"={SETUP_SHEET}!C{SETUP['PRIN']}", f"={WORKUP_SHEET}!C{WORKUP['NEWPRIN']}")
r = sumrow(ws5, r, "Date Granted", f"={SETUP_SHEET}!C{SETUP['GRANT']}", f"={WORKUP_SHEET}!C{WORKUP['NEWGRANT']}", fmt=DATEFMT)
r = sumrow(ws5, r, "Annual Interest Rate", f"={SETUP_SHEET}!C{SETUP['APR']}", f"={WORKUP_SHEET}!C{WORKUP['NEWAPR']}", fmt=PCT2)
r = sumrow(ws5, r, "Number of Installments", f"={SETUP_SHEET}!C{SETUP['NPER']}", f"={WORKUP_SHEET}!C{WORKUP['NEWNPER']}", fmt="0")
r = sumrow(ws5, r, "First Due Date", f"={SETUP_SHEET}!C{SETUP['FIRSTDUE']}", f"={WORKUP_SHEET}!C{WORKUP['NEWFIRSTDUE']}", fmt=DATEFMT)
r = sumrow(ws5, r, "Maturity Date", f"={SETUP_SHEET}!C{SETUP['MATURITY']}", f"={WORKUP_SHEET}!C{WORKUP['NEWMAT']}", fmt=DATEFMT)
r = sumrow(ws5, r, "Fixed Monthly Installment (PHP)", f"={SETUP_SHEET}!C{SETUP['PMT']}", f"={WORKUP_SHEET}!C{WORKUP['NEWPMT']}")

r += 1
ws5.cell(row=r, column=2, value="DEFAULT / RESTRUCTURING BASIS").font = LABEL_FONT
ws5.cell(row=r, column=2).fill = SECTION_FILL
ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
r += 1
for label, ref in [
    ("Last Payment Date", WORKUP['LASTPAY']),
    ("As-of / Report Date", WORKUP['ASOF']),
    ("Days Past Due (DPD)", WORKUP['DPD']),
    ("Outstanding Principal at Default (PHP)", WORKUP['OSPRIN']),
    ("Accrued Interest during Delinquency (PHP)", WORKUP['ACCINT']),
    ("Penalty / Late Fees (PHP)", WORKUP['PENALTY']),
    ("Total Outstanding Exposure (PHP)", WORKUP['TOTEXP']),
    ("Total Waived (not capitalized) (PHP)", WORKUP['WAIVED']),
    ("New Restructured Principal (PHP)", WORKUP['NEWPRIN']),
]:
    fmt = CUR
    if label in ("Last Payment Date", "As-of / Report Date"):
        fmt = DATEFMT
    elif label == "Days Past Due (DPD)":
        fmt = "0"
    ws5.cell(row=r, column=2, value=label).font = LABEL_FONT
    ws5.cell(row=r, column=2).border = BORDER
    c = ws5.cell(row=r, column=3, value=f"={WORKUP_SHEET}!C{ref}")
    c.font = LINK_FONT
    c.number_format = fmt
    c.border = BORDER
    ws5.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    r += 1

# Reorder sheets nicely
order = ["Instructions", "1. Original Loan Setup", "2. Original Amort Schedule",
         "3. Restructuring Workup", "4. Restructured Amort Sched", "5. Summary"]
wb._sheets = [wb[name] for name in order]
tab_colors = [TONIK_DARK, TONIK_PURPLE, TONIK_PURPLE, TONIK_DARK, TONIK_DARK, TONIK_PURPLE]
for name, color in zip(order, tab_colors):
    wsx = wb[name]
    wsx.sheet_view.showGridLines = False
    wsx.sheet_properties.tabColor = color

# freeze header rows on the amortization schedules
ws2.freeze_panes = "A5"
ws4.freeze_panes = "A5"
wb.active = 1  # open on the Loan Setup sheet by default

wb.save("tonik-loan-restructuring-calculator.xlsx")
print("DONE")
